import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Path to your existing results file
excel_path = "/home/devanshu/Downloads/zero-shot-anomaly/AnomalyCLIP/results/full_report/anomaly_results.xlsx"

df = pd.read_excel(excel_path, sheet_name="Results")

rows = []
totals = {'tp': 0, 'tn': 0, 'fp': 0, 'fn': 0}

for cls in sorted(df['class'].unique()):
    sub = df[df['class'] == cls]
    tp = ((sub.true_label == 'Anomalous') & (sub.predicted_label == 'Anomalous')).sum()
    tn = ((sub.true_label == 'Normal') & (sub.predicted_label == 'Normal')).sum()
    fp = ((sub.true_label == 'Normal') & (sub.predicted_label == 'Anomalous')).sum()
    fn = ((sub.true_label == 'Anomalous') & (sub.predicted_label == 'Normal')).sum()
    n = tp + tn + fp + fn
    acc = (tp + tn) / n * 100 if n else 0
    prec = tp / (tp + fp) * 100 if (tp + fp) else 0
    rec = tp / (tp + fn) * 100 if (tp + fn) else 0
    f1 = 2 * prec * rec / (prec + rec) if (prec + rec) else 0
    rows.append({
        'class': cls, 'n': n, 'TP': tp, 'TN': tn, 'FP': fp, 'FN': fn,
        'Accuracy_%': round(acc, 1), 'Precision_%': round(prec, 1),
        'Recall_%': round(rec, 1), 'F1_%': round(f1, 1)
    })
    for k, v in zip(['tp', 'tn', 'fp', 'fn'], [tp, tn, fp, fn]):
        totals[k] += v

n = sum(totals.values())
acc = (totals['tp'] + totals['tn']) / n * 100 if n else 0
prec = totals['tp'] / (totals['tp'] + totals['fp']) * 100 if (totals['tp'] + totals['fp']) else 0
rec = totals['tp'] / (totals['tp'] + totals['fn']) * 100 if (totals['tp'] + totals['fn']) else 0
f1 = 2 * prec * rec / (prec + rec) if (prec + rec) else 0
rows.append({
    'class': 'OVERALL', 'n': n, 'TP': totals['tp'], 'TN': totals['tn'],
    'FP': totals['fp'], 'FN': totals['fn'],
    'Accuracy_%': round(acc, 1), 'Precision_%': round(prec, 1),
    'Recall_%': round(rec, 1), 'F1_%': round(f1, 1)
})

metrics_df = pd.DataFrame(rows)

# Print to terminal
pd.set_option('display.width', 140)
print(metrics_df.to_string(index=False))

# Also write into the existing workbook as a new sheet
wb = load_workbook(excel_path)
if "Metrics" in wb.sheetnames:
    del wb["Metrics"]
ws = wb.create_sheet("Metrics")

ws.append(list(metrics_df.columns))
for _, r in metrics_df.iterrows():
    ws.append(list(r))

header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
for col_idx, col_name in enumerate(metrics_df.columns, start=1):
    cell = ws.cell(row=1, column=col_idx)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions[get_column_letter(col_idx)].width = 14

# Bold the OVERALL row
last_row = len(metrics_df) + 1
for col_idx in range(1, len(metrics_df.columns) + 1):
    ws.cell(row=last_row, column=col_idx).font = Font(bold=True)

ws.freeze_panes = "A2"
wb.save(excel_path)
print(f"\nMetrics sheet added to: {excel_path}")