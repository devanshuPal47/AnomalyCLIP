import argparse
import os
import torch
import numpy as np
import pandas as pd
from scipy.ndimage import gaussian_filter
from skimage import measure
import AnomalyCLIP_lib
from prompt_ensemble import AnomalyCLIP_PromptLearner
from dataset import Dataset
from utils import get_transform, normalize
from visualization import visualizer
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule


def get_bounding_boxes(anomaly_map_norm, threshold, min_area):
    binary_mask = (anomaly_map_norm >= threshold).astype(np.uint8)
    labeled = measure.label(binary_mask)
    boxes = []
    for region in measure.regionprops(labeled):
        if region.area < min_area:
            continue
        min_row, min_col, max_row, max_col = region.bbox
        boxes.append((min_col, min_row, max_col, max_row))  # x1, y1, x2, y2
    boxes.sort(key=lambda b: (b[2] - b[0]) * (b[3] - b[1]), reverse=True)
    return boxes


def main(args):
    device = "cuda" if torch.cuda.is_available() else "cpu"
    print(f"Using device: {device}")

    AnomalyCLIP_parameters = {
        "Prompt_length": args.n_ctx,
        "learnabel_text_embedding_depth": args.depth,
        "learnabel_text_embedding_length": args.t_n_ctx
    }
    model, _ = AnomalyCLIP_lib.load("ViT-L/14@336px", device=device, design_details=AnomalyCLIP_parameters)
    model.eval()

    preprocess, target_transform = get_transform(args)
    test_data = Dataset(root=args.data_path, transform=preprocess, target_transform=target_transform, dataset_name=args.dataset)
    test_dataloader = torch.utils.data.DataLoader(test_data, batch_size=1, shuffle=False)

    prompt_learner = AnomalyCLIP_PromptLearner(model.to("cpu"), AnomalyCLIP_parameters)
    checkpoint = torch.load(args.checkpoint_path)
    prompt_learner.load_state_dict(checkpoint["prompt_learner"])
    prompt_learner.to(device)
    model.to(device)
    model.visual.DAPM_replace(DPAM_layer=20)

    prompts, tokenized_prompts, compound_prompts_text = prompt_learner(cls_id=None)
    text_features = model.encode_text_learn(prompts, tokenized_prompts, compound_prompts_text).float()
    text_features = torch.stack(torch.chunk(text_features, dim=0, chunks=2), dim=1)
    text_features = text_features / text_features.norm(dim=-1, keepdim=True)
    model.to(device)

    rows = []
    total = len(test_dataloader)

    for i, items in enumerate(test_dataloader):
        cls_name = items['cls_name'][0]
        img_path = items['img_path'][0]
        image = items['img'].to(device)

        try:
            true_anomaly = int(items['anomaly'][0])
        except Exception:
            defect_type_from_path = img_path.split('/')[-2]
            true_anomaly = 0 if defect_type_from_path == 'good' else 1

        with torch.no_grad():
            image_features, patch_features = model.encode_image(image, args.features_list, DPAM_layer=20)
            image_features = image_features / image_features.norm(dim=-1, keepdim=True)

            # image-level normal-vs-anomalous score
            sim_img = image_features @ text_features[0].permute(1, 0)  # (1,2): [normal, abnormal]
            image_score = ((sim_img[:, 1] + 1 - sim_img[:, 0]) / 2).item()

            # pixel-level anomaly map (for localization + heatmap)
            anomaly_map_list = []
            for idx, patch_feature in enumerate(patch_features):
                if idx >= args.feature_map_layer[0]:
                    patch_feature = patch_feature / patch_feature.norm(dim=-1, keepdim=True)
                    similarity, _ = AnomalyCLIP_lib.compute_similarity(patch_feature, text_features[0])
                    similarity_map = AnomalyCLIP_lib.get_similarity_map(similarity[:, 1:, :], args.image_size)
                    a_map = (similarity_map[..., 1] + 1 - similarity_map[..., 0]) / 2.0
                    anomaly_map_list.append(a_map)
            anomaly_map = torch.stack(anomaly_map_list).sum(dim=0)
            anomaly_map = torch.stack(
                [torch.from_numpy(gaussian_filter(m, sigma=args.sigma)) for m in anomaly_map.detach().cpu()],
                dim=0
            )

        anomaly_map_full = anomaly_map.detach().cpu().numpy()  # shape (1, H, W)
        anomaly_map_np = anomaly_map_full[0]
        anomaly_map_norm = normalize(anomaly_map_np)

        boxes = get_bounding_boxes(anomaly_map_norm, args.threshold, args.min_region_area)
        area_pct = float((anomaly_map_norm >= args.threshold).sum()) / anomaly_map_norm.size * 100

        predicted_label = "Anomalous" if image_score >= args.classify_threshold else "Normal"
        true_label = "Anomalous" if true_anomaly == 1 else "Normal"
        loc_str = "; ".join([f"({x1},{y1})-({x2},{y2})" for x1, y1, x2, y2 in boxes[:3]]) if boxes else ""

        heatmap_path = ""
        if args.save_heatmaps:
            visualizer([img_path], anomaly_map_full, args.image_size, args.save_path, [cls_name])
            defect_type = img_path.split('/')[-2]
            filename = img_path.split('/')[-1]
            heatmap_path = os.path.join(args.save_path, 'imgs', cls_name, defect_type, filename)

        rows.append({
            "image_path": img_path,
            "class": cls_name,
            "true_label": true_label,
            "predicted_label": predicted_label,
            "anomaly_score": round(image_score, 4),
            "defect_area_pct": round(area_pct, 2),
            "damage_location_bbox_xyxy": loc_str,
            "heatmap_path": heatmap_path,
        })

        if (i + 1) % 50 == 0 or (i + 1) == total:
            print(f"Processed {i+1}/{total}")

    df = pd.DataFrame(rows)
    os.makedirs(args.save_path, exist_ok=True)
    excel_path = os.path.join(args.save_path, "anomaly_results.xlsx")
    df.to_excel(excel_path, index=False, sheet_name="Results")

    wb = load_workbook(excel_path)
    ws = wb["Results"]
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        col_letter = get_column_letter(col_idx)
        max_len = max(df[col_name].astype(str).map(len).max() if len(df) else 0, len(col_name))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)
    ws.freeze_panes = "A2"

    pred_col_letter = get_column_letter(df.columns.get_loc("predicted_label") + 1)
    last_row = len(df) + 1
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    rng = f"{pred_col_letter}2:{pred_col_letter}{last_row}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Anomalous"'], fill=red_fill))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Normal"'], fill=green_fill))

    tp = ((df.true_label == "Anomalous") & (df.predicted_label == "Anomalous")).sum()
    tn = ((df.true_label == "Normal") & (df.predicted_label == "Normal")).sum()
    fp = ((df.true_label == "Normal") & (df.predicted_label == "Anomalous")).sum()
    fn = ((df.true_label == "Anomalous") & (df.predicted_label == "Normal")).sum()
    acc = (tp + tn) / len(df) * 100 if len(df) else 0
    summary_ws = wb.create_sheet("Summary")
    for r in [
        ["Total images", len(df)],
        ["True Positives (correctly flagged anomalous)", int(tp)],
        ["True Negatives (correctly flagged normal)", int(tn)],
        ["False Positives (normal flagged as anomalous)", int(fp)],
        ["False Negatives (anomalous missed)", int(fn)],
        ["Accuracy %", round(acc, 2)],
    ]:
        summary_ws.append(r)
    summary_ws.column_dimensions["A"].width = 48
    summary_ws.column_dimensions["B"].width = 15

    wb.save(excel_path)
    print(f"\nDone. Excel report: {excel_path}")
    if args.save_heatmaps:
        print(f"Heatmaps: {os.path.join(args.save_path, 'imgs')}")


if __name__ == '__main__':
    parser = argparse.ArgumentParser("AnomalyCLIP classification + localization report")
    parser.add_argument("--data_path", type=str, default="./data/mvtec_ad")
    parser.add_argument("--save_path", type=str, default="./results/full_report")
    parser.add_argument("--checkpoint_path", type=str, default="./checkpoints/9_12_4_multiscale_visa/epoch_15.pth")
    parser.add_argument("--dataset", type=str, default="mvtec")
    parser.add_argument("--features_list", type=int, nargs="+", default=[6, 12, 18, 24])
    parser.add_argument("--image_size", type=int, default=518)
    parser.add_argument("--depth", type=int, default=9)
    parser.add_argument("--n_ctx", type=int, default=12)
    parser.add_argument("--t_n_ctx", type=int, default=4)
    parser.add_argument("--feature_map_layer", type=int, nargs="+", default=[0, 1, 2, 3])
    parser.add_argument("--sigma", type=int, default=4)
    parser.add_argument("--threshold", type=float, default=0.5, help="pixel threshold (0-1) for marking a pixel as damaged")
    parser.add_argument("--classify_threshold", type=float, default=0.5, help="image-level threshold (0-1) for Normal vs Anomalous")
    parser.add_argument("--min_region_area", type=int, default=50, help="ignore defect blobs smaller than this many pixels (noise filter)")
    parser.add_argument("--save_heatmaps", action="store_true", default=True)
    parser.add_argument("--no_heatmaps", dest="save_heatmaps", action="store_false")
    args = parser.parse_args()
    main(args)